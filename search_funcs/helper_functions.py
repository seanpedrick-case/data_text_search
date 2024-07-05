import os
import re
import pandas as pd
import gradio as gr
import os
import shutil
import getpass
import gzip
import zipfile
import pickle
import numpy as np

from typing import List

# Openpyxl functions for output
from openpyxl import Workbook
from openpyxl.cell.text import InlineFont 
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.styles import Font, Alignment

from search_funcs.aws_functions import bucket_name

megabyte = 1024 * 1024  # Bytes in a megabyte
file_size_mb = 500  # Size in megabytes
file_size_bytes_500mb =  megabyte * file_size_mb

def get_or_create_env_var(var_name, default_value):
    # Get the environment variable if it exists
    value = os.environ.get(var_name)
    
    # If it doesn't exist, set it to the default value
    if value is None:
        os.environ[var_name] = default_value
        value = default_value
    
    return value

# Retrieving or setting output folder
output_folder = get_or_create_env_var('GRADIO_OUTPUT_FOLDER', 'output/')
print(f'The value of GRADIO_OUTPUT_FOLDER is {output_folder}')

# Retrieving or setting RUNNING_ON_APP_RUNNER
running_on_app_runner_var = get_or_create_env_var('RUNNING_ON_APP_RUNNER', '0')
print(f'The value of RUNNING_ON_APP_RUNNER is {running_on_app_runner_var}')



def ensure_output_folder_exists(output_folder):
    """Checks if the output folder exists, creates it if not."""

    folder_name = output_folder

    if not os.path.exists(folder_name):
        # Create the folder if it doesn't exist
        os.makedirs(folder_name)
        print(f"Created the output folder:", folder_name)
    else:
        print(f"The output folder already exists:", folder_name)

def get_connection_params(request: gr.Request):
        if request:
            #request_data = request.json()  # Parse JSON body
            #print("All request data:", request_data)
            #context_value = request_data.get('context') 
            #if 'context' in request_data:
            #     print("Request context dictionary:", request_data['context'])

            #print("Request headers dictionary:", request.headers)
            #print("All host elements", request.client)           
            #print("IP address:", request.client.host)
            #print("Query parameters:", dict(request.query_params))
            # To get the underlying FastAPI items you would need to use await and some fancy @ stuff for a live query: https://fastapi.tiangolo.com/vi/reference/request/
            #print("Request dictionary to object:", request.request.body())
            #print("Session hash:", request.session_hash)

            # Retrieving or setting CUSTOM_CLOUDFRONT_HEADER
            CUSTOM_CLOUDFRONT_HEADER_var = get_or_create_env_var('CUSTOM_CLOUDFRONT_HEADER', '')
            print(f'The value of CUSTOM_CLOUDFRONT_HEADER is {CUSTOM_CLOUDFRONT_HEADER_var}')

            # Retrieving or setting CUSTOM_CLOUDFRONT_HEADER_VALUE
            CUSTOM_CLOUDFRONT_HEADER_VALUE_var = get_or_create_env_var('CUSTOM_CLOUDFRONT_HEADER_VALUE', '')
            print(f'The value of CUSTOM_CLOUDFRONT_HEADER_VALUE_var is {CUSTOM_CLOUDFRONT_HEADER_VALUE_var}')

            if CUSTOM_CLOUDFRONT_HEADER_var and CUSTOM_CLOUDFRONT_HEADER_VALUE_var:
                if CUSTOM_CLOUDFRONT_HEADER_var in request.headers:
                    supplied_cloudfront_custom_value = request.headers[CUSTOM_CLOUDFRONT_HEADER_var]
                    if supplied_cloudfront_custom_value == CUSTOM_CLOUDFRONT_HEADER_VALUE_var:
                        print("Custom Cloudfront header found:", supplied_cloudfront_custom_value)
                    else:
                        raise(ValueError, "Custom Cloudfront header value does not match expected value.")

            if 'x-cognito-id' in request.headers:
                out_session_hash = request.headers['x-cognito-id']
                base_folder = "user-files/"
                print("Cognito ID found:", out_session_hash)

            else:
                out_session_hash = request.session_hash
                base_folder = "temp-files/"
                #print("Cognito ID not found. Using session hash as save folder.")

            output_folder = base_folder + out_session_hash + "/"
            #if bucket_name:
            #    print("S3 output folder is: " + "s3://" + bucket_name + "/" + output_folder)

            return out_session_hash, output_folder
        else:
            print("No session parameters found.")
            return "", ""

# Attempt to delete content of gradio temp folder
def get_temp_folder_path():
    username = getpass.getuser()
    return os.path.join('C:\\Users', username, 'AppData\\Local\\Temp\\gradio')

def empty_folder(directory_path):
    if not os.path.exists(directory_path):
        #print(f"The directory {directory_path} does not exist. No temporary files from previous app use found to delete.")
        return

    for filename in os.listdir(directory_path):
        file_path = os.path.join(directory_path, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            #print(f'Failed to delete {file_path}. Reason: {e}')
            print('')

def get_file_path_end(file_path):
    # First, get the basename of the file (e.g., "example.txt" from "/path/to/example.txt")
    basename = os.path.basename(file_path)
    
    # Then, split the basename and its extension and return only the basename without the extension
    filename_without_extension, _ = os.path.splitext(basename)

    #print(filename_without_extension)
    
    return filename_without_extension

def get_file_path_end_with_ext(file_path):
    match = re.search(r'(.*[\/\\])?(.+)$', file_path)
        
    filename_end = match.group(2) if match else ''

    return filename_end

def ensure_output_folder_exists(output_folder):
    """Checks if the output folder exists, creates it if not."""

    folder_name = output_folder

    if not os.path.exists(folder_name):
        # Create the folder if it doesn't exist
        os.makedirs(folder_name)
        print(f"Created the output folder:", folder_name)
    else:
        print(f"The output folder already exists:", folder_name)

def detect_file_type(filename):
    """Detect the file type based on its extension."""
    if (filename.endswith('.csv')) | (filename.endswith('.csv.gz')) | (filename.endswith('.zip')):
        return 'csv'
    elif filename.endswith('.xlsx'):
        return 'xlsx'
    elif filename.endswith('.parquet'):
        return 'parquet'
    elif filename.endswith('.pkl.gz'):
        return 'pkl.gz'
    #elif filename.endswith('.gz'):
    #    return 'gz'
    else:
        raise ValueError("Unsupported file type.")

def read_file(filename):
    """Read the file based on its detected type."""
    file_type = detect_file_type(filename)
        
    print("Loading in file")

    if file_type == 'csv':
        file = pd.read_csv(filename, low_memory=False).reset_index()#.drop(["index", "Unnamed: 0"], axis=1, errors="ignore")
    elif file_type == 'xlsx':
        file = pd.read_excel(filename).reset_index()#.drop(["index", "Unnamed: 0"], axis=1, errors="ignore")
    elif file_type == 'parquet':
        file = pd.read_parquet(filename).reset_index()#.drop(["index", "Unnamed: 0"], axis=1, errors="ignore")
    elif file_type == 'pkl.gz':
        with gzip.open(filename, 'rb') as file:
            file = pickle.load(file)
    #elif file_type == ".gz":
    #    with gzip.open(filename, 'rb') as file:
    #        file = pickle.load(file)

    print("File load complete")

    return file

def process_zip_files(file_list, progress=gr.Progress(track_tqdm=True)):
    """
    Processes a list of file names, unzipping any ZIP files found
    and adding the extracted file names to the list.

    Args:
        file_list: A list of file names (strings).
    """
    progress(0.1, desc="Unzipping zip files")

    i = 0
    while i < len(file_list):  # Use 'while' for dynamic list changes
        file_path = file_list[i]

        if file_path.endswith(".zip"):
            try:
                zip_dir = os.path.dirname(file_path) or "."  # Get zip file's directory or use current if none
                with zipfile.ZipFile(file_path, 'r') as zip_ref:
                    zip_ref.extractall(zip_dir)  # Extract to zip's directory
                    #print("List of files in zip:", zip_ref.namelist())
                    extracted_files = [os.path.join(zip_dir, name) for name in zip_ref.namelist()]  
                    file_list.extend(extracted_files)
                    
            except zipfile.BadZipFile:
                print(f"Warning: '{file_path}' is not a valid zip file.")

        i += 1

    file_list = [file for file in file_list if not file.endswith(".zip")]
    print("file_list after files in zip extracted:", file_list)

    return file_list

def initial_data_load(in_file:List[str], progress = gr.Progress(track_tqdm=True)):
    '''
    When file is loaded, update the column dropdown choices and relevant state variables
    '''
    new_choices = []
    concat_choices = []
    index_load = None
    embed_load = np.array([])
    tokenised_load = []
    out_message = ""
    current_source = ""
    df = pd.DataFrame()

    file_list = [string.name for string in in_file]

    # If a zip file is loaded, unzip it and add the file names to the file_list
    file_list = process_zip_files(file_list)

    #print("File_list that makes it to main data load function:", file_list)         

    progress(0.3, desc="Loading in data files")

    data_file_names = [string for string in file_list if "tokenised" not in string.lower() and "npz" not in string.lower() and "search_index" not in string.lower()]
    print("Data file names:", data_file_names)

    if not data_file_names:
        out_message = "Please load in at least one csv/Excel/parquet data file."
        print(out_message)
        return gr.Dropdown(choices=concat_choices), gr.Dropdown(choices=concat_choices), pd.DataFrame(), pd.DataFrame(), index_load, embed_load, tokenised_load, out_message, None

    # This if you have loaded in a documents object for the semantic search
    if "pkl" in data_file_names[0]: 
        print("Document object for semantic search:", data_file_names[0])
        df = read_file(data_file_names[0])
        new_choices = list(df[0].metadata.keys()) #["Documents"] #["page_contents"] + 
        current_source = get_file_path_end_with_ext(data_file_names[0])  

    # This if you have loaded in a csv/parquets/xlsx
    else:
        for file in data_file_names:

            current_source = current_source + get_file_path_end_with_ext(file) + " "
        
            # Get the size of the file
            print("Checking file size")
            file_size = os.path.getsize(file)
            if file_size > file_size_bytes_500mb:
                out_message = "Data file greater than 500mb in size. Please use smaller sizes."
                print(out_message)
                return gr.Dropdown(choices=concat_choices), gr.Dropdown(choices=concat_choices), pd.DataFrame(), pd.DataFrame(), index_load, embed_load, tokenised_load, out_message, None


            df_new = read_file(file)

            df = pd.concat([df, df_new], ignore_index = True)

        new_choices = list(df.columns)

    concat_choices.extend(new_choices)

    progress(0.6, desc="Loading in embedding/search index files")

    # Check if there is a search index file already
    index_file_names = [string for string in file_list if ".gz" in string.lower()]

    if index_file_names:
        index_file_name = index_file_names[0]
        print("Search index file name found:", index_file_name)
        index_load = read_file(index_file_name)

    embeddings_file_names = [string for string in file_list if "embedding" in string.lower()]

    if embeddings_file_names:
        print("Loading embeddings from file.")
        embed_load = np.load(embeddings_file_names[0])['arr_0']

        # If embedding files have 'super_compress' in the title, they have been multiplied by 100 before save
        if "compress" in embeddings_file_names[0]:
            embed_load /= 100
    else:
        embed_load = np.array([])

    tokenised_file_names = [string for string in file_list if "tokenised" in string.lower()]
    if tokenised_file_names:
        tokenised_load = read_file(tokenised_file_names[0])

    out_message = "Initial data load successful. Next, choose a data column to search in the drop down above, then click 'Load data'"
    print(out_message)
        
    return gr.Dropdown(choices=concat_choices), gr.Dropdown(choices=concat_choices), df, df, index_load, embed_load, tokenised_load, out_message, current_source, file_list

def put_columns_in_join_df(in_file:str):
    '''
    When file is loaded, update the column dropdown choices
    '''
    new_df = pd.DataFrame()
    #print("in_bm25_column")

    new_choices = []
    concat_choices = []
    
    
    new_df = read_file(in_file.name)
    new_choices = list(new_df.columns)

    #print(new_choices)

    concat_choices.extend(new_choices)

    out_message = "File load successful. Now select a column to join below."    
        
    return gr.Dropdown(choices=concat_choices), new_df, out_message

def load_spacy_model():
	# Load the SpaCy model
	from spacy.cli.download import download
	import spacy
	spacy.prefer_gpu()

	try:
		import en_core_web_sm
		nlp = en_core_web_sm.load()
		print("Successfully imported spaCy model")
	except:
		download("en_core_web_sm")
		nlp = spacy.load("en_core_web_sm")
		print("Successfully imported spaCy model")
	return nlp

def display_info(info_component):
    gr.Info(info_component)

def highlight_found_text(search_text: str, full_text: str) -> str:
    """
    Highlights occurrences of search_text within full_text.
    
    Parameters:
    - search_text (str): The text to be searched for within full_text.
    - full_text (str): The text within which search_text occurrences will be highlighted.
    
    Returns:
    - str: A string with occurrences of search_text highlighted.
    
    Example:
    >>> highlight_found_text("world", "Hello, world! This is a test. Another world awaits.")
    'Hello, <mark style="color:black;">world</mark>! This is a test. Another <mark style="color:black;">world</mark> awaits.'
    """

    def extract_text_from_input(text, i=0):
        if isinstance(text, str):
            return text
        elif isinstance(text, list):
            return text[i][0]
        else:
            return ""

    def extract_search_text_from_input(text):
        if isinstance(text, str):
            return text
        elif isinstance(text, list):
            return text[-1][1]
        else:
            return ""

    full_text = extract_text_from_input(full_text)
    search_text = extract_search_text_from_input(search_text)

    sections = search_text.split(sep = " ")

    found_positions = {}
    for x in sections:
        text_start_pos = 0
        while text_start_pos != -1:
            text_start_pos = full_text.find(x, text_start_pos)
            if text_start_pos != -1:
                found_positions[text_start_pos] = text_start_pos + len(x)
                text_start_pos += 1

    # Combine overlapping or adjacent positions
    sorted_starts = sorted(found_positions.keys())
    combined_positions = []
    if sorted_starts:
        current_start, current_end = sorted_starts[0], found_positions[sorted_starts[0]]
        for start in sorted_starts[1:]:
            if start <= (current_end + 10):
                current_end = max(current_end, found_positions[start])
            else:
                combined_positions.append((current_start, current_end))
                current_start, current_end = start, found_positions[start]
        combined_positions.append((current_start, current_end))

    # Construct pos_tokens
    pos_tokens = []
    prev_end = 0
    for start, end in combined_positions:
        if end-start > 1: # Only combine if there is a significant amount of matched text. Avoids picking up single words like 'and' etc.
            pos_tokens.append(full_text[prev_end:start])
            pos_tokens.append('<mark style="color:black;">' + full_text[start:end] + '</mark>')
            prev_end = end
    pos_tokens.append(full_text[prev_end:])

    return "".join(pos_tokens), combined_positions

def create_rich_text_cell_from_positions(full_text: str, combined_positions: list[tuple[int, int]]) -> CellRichText:
    """
    Create a rich text cell with highlighted positions.

    This function takes the full text and a list of combined positions, and creates a rich text cell
    with the specified positions highlighted in red.

    Parameters:
    full_text (str): The full text to be processed.
    combined_positions (list[tuple[int, int]]): A list of tuples representing the start and end positions to be highlighted.

    Returns:
    CellRichText: The created rich text cell with highlighted positions.
    """
    # Construct pos_tokens
    red = InlineFont(color='00FF0000')
    rich_text_cell = CellRichText()

    prev_end = 0
    for start, end in combined_positions:
        if end-start > 1: # Only combine if there is a significant amount of matched text. Avoids picking up single words like 'and' etc.
            rich_text_cell.append(full_text[prev_end:start])
            rich_text_cell.append(TextBlock(red, full_text[start:end]))
            prev_end = end
    rich_text_cell.append(full_text[prev_end:])

    return rich_text_cell

def create_highlighted_excel_wb(df: pd.DataFrame, search_text: str, column_to_highlight: str) -> Workbook:
    """
    Create a new Excel workbook with highlighted search text.

    This function takes a DataFrame, a search text, and a column name to highlight. It creates a new Excel workbook,
    highlights the occurrences of the search text in the specified column, and returns the workbook.

    Parameters:
    df (pd.DataFrame): The DataFrame containing the data to be written to the Excel workbook.
    search_text (str): The text to search for and highlight in the specified column.
    column_to_highlight (str): The name of the column in which to highlight the search text.

    Returns:
    Workbook: The created Excel workbook with highlighted search text.
    """

    # Create a new Excel workbook
    wb = Workbook()
    sheet = wb.active

    # Insert headers into the worksheet, make bold
    sheet.append(df.columns.tolist())
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    column_width = 150  # Adjust as needed
    relevant_column_no = (df.columns == column_to_highlight).argmax() + 1
    print(relevant_column_no)
    sheet.column_dimensions[sheet.cell(row=1, column=relevant_column_no).column_letter].width = column_width

    # Find substrings in cells and highlight
    for r_idx, row in enumerate(df.itertuples(), start=2):
        for c_idx, cell_value in enumerate(row[1:], start=1):
            sheet.cell(row=r_idx, column=c_idx, value=cell_value)
            if df.columns[c_idx - 1] == column_to_highlight:

                html_text, combined_positions = highlight_found_text(search_text, cell_value)
                sheet.cell(row=r_idx, column=c_idx).value = create_rich_text_cell_from_positions(cell_value, combined_positions)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(wrap_text=True)

    return wb